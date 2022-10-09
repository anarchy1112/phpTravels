import logging

import openpyxl
from utilities.logger import Loger

log=Loger(__name__,logging.INFO)


def data_getter(sheetname):
    book=openpyxl.load_workbook('..//excel//Book1.xlsx')
    sheet=book[sheetname]
    total_rows=sheet.max_row
    total_cols=sheet.max_column

    main=[]
    for i in range(2, total_rows+1):
   #     log.logger.debug('inside loop')
        temp=[]
        for j in range(1,total_cols+1):
            temp.insert(j, sheet.cell(row=i,column=j).value)
          #  log.logger.info('end of loop')
        main.insert(i, temp)
    return main


